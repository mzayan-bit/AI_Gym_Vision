import streamlit as st
import cv2
import numpy as np
import mediapipe as mp
import pandas as pd
import time
from datetime import datetime
import tempfile
import os
import openpyxl
import matplotlib.pyplot as plt
import requests
from collections import deque
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
from enum import Enum
from ultralytics import YOLO
import yaml
import seaborn as sns
st.set_page_config(page_title="AI Gym Vision", layout="wide", page_icon="🏋️")
def load_lottie_url(url: str):
    try:
        r = requests.get(url, timeout=5)
        if r.status_code != 200: return None
        return r.json()
    except: return None
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
GYM_EQUIPMENT_MODEL_PATH = os.path.join(BASE_DIR, 'trained_model', 'gym.pt')
GYM_EQUIPMENT_DATA_YAML_PATH = os.path.join(BASE_DIR, 'trained_model', 'gym_inference_data.yaml')
PERSON_DETECTION_MODEL_PATH = os.path.join(BASE_DIR, 'yolov8n.pt')

MIN_VISIBILITY_THRESHOLD = 0.60
POSTURE_EVALUATION_FRAMES = 15
class PostureQuality(Enum):
    EXCELLENT = "EXCELLENT"
    GOOD = "GOOD"
    AVERAGE = "AVERAGE"
    POOR = "POOR"
    VERY_POOR = "VERY_POOR"

@dataclass
class PostureMetrics:
    angle_consistency: float = 0.0
    range_of_motion: float = 0.0
    form_stability: float = 0.0
    timing_consistency: float = 0.0
    tempo_score: float = 0.0
    overall_score: float = 0.0
    quality: PostureQuality = PostureQuality.AVERAGE

class EnhancedExerciseState:
    def __init__(self, person_id: int):
        self.person_id = person_id
        self.exercise_name = "Free Movement"
        self.reps = 0
        self.stage = "start"
        self.feedback = "Initializing analysis..."
        self.angle_history = deque(maxlen=20)
        self.smoothed_angles = deque(maxlen=POSTURE_EVALUATION_FRAMES)
        self.min_angle_in_rep = 180
        self.max_angle_in_rep = 0
        self.rep_start_time = time.time()
        self.phase_start_time = time.time()
        self.eccentric_duration = 0
        self.concentric_duration = 0
        self.rep_durations = deque(maxlen=10)
        self.posture_metrics = PostureMetrics()
        self.form_violations = []
        self.stability_scores = deque(maxlen=20)
        self.quality_history = []
        self.form_score_history = []
        self.velocity_history = deque(maxlen=5)
        self.acceleration_history = deque(maxlen=5)
        self.previous_angle = None
        self.previous_velocity = 0
        self.hold_duration = 0
        
    def update_angle_data(self, current_angle: float, smoothing_factor: int):
        self.angle_history.append(current_angle)
        if len(self.angle_history) >= smoothing_factor:
            smoothing_window = list(self.angle_history)[-smoothing_factor:]
            smoothed_angle = np.mean(smoothing_window)
            self.smoothed_angles.append(smoothed_angle)
            if self.previous_angle is not None:
                velocity = smoothed_angle - self.previous_angle
                self.velocity_history.append(abs(velocity))
                acceleration = velocity - self.previous_velocity
                self.acceleration_history.append(abs(acceleration))
                self.previous_velocity = velocity
            self.previous_angle = smoothed_angle
            return smoothed_angle
        return current_angle
    
    def evaluate_posture_quality(self, exercise_rules: Dict) -> PostureMetrics:
        if len(self.smoothed_angles) < POSTURE_EVALUATION_FRAMES:
            return self.posture_metrics
        angles = list(self.smoothed_angles)
        angle_std = np.std(angles[-10:]) if len(angles) >= 10 else np.std(angles)
        consistency_score = max(0, 100 - (angle_std * 2))
        current_rom = self.max_angle_in_rep - self.min_angle_in_rep
        target_rom = exercise_rules.get('target_rom', exercise_rules['min_movement_diff'])
        rom_score = min(100, (current_rom / target_rom) * 100)
        
        if len(self.velocity_history) >= 3:
            velocity_consistency = 100 - (np.std(list(self.velocity_history)) * 10)
            stability_score = max(0, min(100, velocity_consistency))
        else:
            stability_score = 50
            
        if len(self.rep_durations) >= 3:
            duration_std = np.std(list(self.rep_durations))
            timing_score = max(0, 100 - (duration_std * 20))
        else:
            timing_score = 50
            
        ideal_tempo = exercise_rules.get('ideal_tempo')
        tempo_score = 50
        if ideal_tempo and self.eccentric_duration > 0 and self.concentric_duration > 0:
            ideal_ecc = ideal_tempo.get('eccentric', 2.0)
            ideal_con = ideal_tempo.get('concentric', 1.5)
            ecc_error = abs(self.eccentric_duration - ideal_ecc) / ideal_ecc
            con_error = abs(self.concentric_duration - ideal_con) / ideal_con
            total_error = (ecc_error + con_error) / 2
            tempo_score = max(0, 100 * (1 - total_error))

        weights = {'consistency': 0.20, 'rom': 0.30, 'stability': 0.20, 'timing': 0.10, 'tempo': 0.20}
        overall_score = (
            consistency_score * weights['consistency'] +
            rom_score * weights['rom'] +
            stability_score * weights['stability'] +
            timing_score * weights['timing'] +
            tempo_score * weights['tempo']
        )
        
        if overall_score >= 90: quality = PostureQuality.EXCELLENT
        elif overall_score >= 80: quality = PostureQuality.GOOD
        elif overall_score >= 65: quality = PostureQuality.AVERAGE
        elif overall_score >= 50: quality = PostureQuality.POOR
        else: quality = PostureQuality.VERY_POOR

        self.posture_metrics = PostureMetrics(
            angle_consistency=consistency_score, range_of_motion=rom_score,
            form_stability=stability_score, timing_consistency=timing_score,
            tempo_score=tempo_score,
            overall_score=overall_score, quality=quality
        )
        return self.posture_metrics
    
    def reset_rep_cycle(self):
        current_time = time.time()
        rep_duration = current_time - self.rep_start_time
        self.rep_durations.append(rep_duration)
        self.quality_history.append(self.posture_metrics.quality)
        self.stage = "start"
        self.min_angle_in_rep = 180
        self.max_angle_in_rep = 0
        self.rep_start_time = current_time
        self.phase_start_time = current_time
        self.form_violations.clear()

@st.cache_resource
def load_models():
    try:
        person_model = YOLO(PERSON_DETECTION_MODEL_PATH)
        equipment_model = YOLO(GYM_EQUIPMENT_MODEL_PATH)
        return person_model, equipment_model
    except Exception as e:
        st.error(f"❌ Error loading YOLO models: {e}. Make sure model files are in the correct directory.")
        return None, None

@st.cache_data
def load_equipment_classes():
    try:
        with open(GYM_EQUIPMENT_DATA_YAML_PATH, 'r') as f:
            gym_data_yaml = yaml.safe_load(f)
            return gym_data_yaml.get('names', [])
    except Exception as e:
        st.error(f"❌ Error loading {GYM_EQUIPMENT_DATA_YAML_PATH}: {e}")
        return []

ENHANCED_EXERCISE_RULES = {
    'chest press machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 85, 'max_rep_angle': 170, 'min_movement_diff': 50, 'target_rom': 85,
        'ideal_tempo': {'eccentric': 2.0, 'concentric': 1.5},
        'critical_angles': {'danger_zone': (70, 90)},
        'rep_threshold': 0.7, 'smoothing_factor': 3,
        'feedback': {
            'excellent': "Outstanding chest press form! Perfect range and control.",
            'good': "Good form. Maintain steady tempo.", 'average': "Average form. Focus on full range of motion.",
            'poor': "Poor form. Check elbow position and alignment.", 'very_poor': "Dangerous form! Risk of shoulder injury.",
            'down': "Lower with control to chest level.", 'up': "Press with power, full extension.",
            'incomplete': "Partial rep. Ensure full range of motion.",
            'form_violation': "Elbow positioning needs attention!"
        }
    },
    'lat pull down': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 50, 'max_rep_angle': 175, 'min_movement_diff': 70, 'target_rom': 125,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 1.8},
        'critical_angles': {'danger_zone': (40, 55)},
        'rep_threshold': 0.75, 'smoothing_factor': 4,
        'feedback': {
            'excellent': "Perfect lat pulldown! Great range and control.",
            'good': "Good form. Focus on squeezing shoulder blades.", 'average': "Average form. Pull lower to chest level.",
            'poor': "Poor form. Avoid using momentum.", 'very_poor': "Very poor form! Risk of shoulder injury.",
            'down': "Pull down to chest, squeeze shoulder blades together.", 'up': "Control the return, feel the lat stretch.",
            'incomplete': "Incomplete rep. Pull to chest level.",
            'form_violation': "Shoulder position needs correction!"
        }
    },
    
    'arm curl machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 40, 'max_rep_angle': 160, 'min_movement_diff': 70, 'target_rom': 120,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 1.5},
        'critical_angles': {'danger_zone': (30, 45)},
        'rep_threshold': 0.8, 'smoothing_factor': 2,
        'feedback': {
            'excellent': "Perfect bicep curl! Full range with control.",
            'good': "Good form. Maintain steady tempo.", 'average': "Average form. Focus on full range of motion.",
            'poor': "Poor form. Avoid swinging or momentum.", 'very_poor': "Very poor form! Risk of elbow strain.",
            'down': "Lower with control, full extension.", 'up': "Curl up smoothly, squeeze at the top.",
            'incomplete': "Incomplete rep. Ensure full range.",
            'form_violation': "Elbow position needs correction!"
        }
    },
    'chest fly machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 70, 'max_rep_angle': 165, 'min_movement_diff': 60, 'target_rom': 95,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 2.0},
        'critical_angles': {'danger_zone': (55, 75)},
        'rep_threshold': 0.75, 'smoothing_factor': 4,
        'feedback': {
            'excellent': "Outstanding chest fly form! Perfect arc motion.",
            'good': "Good form. Maintain the arc motion.", 'average': "Average form. Focus on chest squeeze.",
            'poor': "Poor form. Keep elbows slightly bent.", 'very_poor': "Dangerous form! Risk of shoulder injury.",
            'down': "Open with control, feel the chest stretch.", 'up': "Squeeze chest muscles, bring arms together.",
            'incomplete': "Incomplete rep. Full range needed.",
            'form_violation': "Arm position needs adjustment!"
        }
    },
    
    'lateral raises machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 20, 'max_rep_angle': 140, 'min_movement_diff': 75, 'target_rom': 120,
        'ideal_tempo': {'eccentric': 2.0, 'concentric': 1.5},
        'critical_angles': {'danger_zone': (10, 25)},
        'rep_threshold': 0.8, 'smoothing_factor': 2,
        'feedback': {
            'excellent': "Perfect lateral raise! Great shoulder activation.",
            'good': "Good form. Control the tempo.", 'average': "Average form. Lift to shoulder height.",
            'poor': "Poor form. Avoid using momentum.", 'very_poor': "Very poor form! Risk of shoulder impingement.",
            'down': "Lower with control, resist gravity.", 'up': "Lift to shoulder height, lead with pinkies.",
            'incomplete': "Incomplete rep. Lift to shoulder level.",
            'form_violation': "Shoulder alignment needs attention!"
        }
    },
    'leg extension': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_HIP, mp.solutions.pose.PoseLandmark.LEFT_KNEE, mp.solutions.pose.PoseLandmark.LEFT_ANKLE],
        'min_rep_angle': 70, 'max_rep_angle': 170, 'min_movement_diff': 65, 'target_rom': 100,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 1.8},
        'critical_angles': {'danger_zone': (55, 75)},
        'rep_threshold': 0.75, 'smoothing_factor': 3,
        'feedback': {
            'excellent': "Excellent leg extension! Perfect quad activation.",
            'good': "Good form. Control the movement.", 'average': "Average form. Focus on full extension.",
            'poor': "Poor form. Avoid bouncing at the top.", 'very_poor': "Dangerous form! Risk of knee hyperextension.",
            'down': "Lower with control, don't drop.", 'up': "Extend smoothly to full extension.",
            'incomplete': "Incomplete rep. Achieve full extension.",
            'form_violation': "Knee alignment needs correction!"
        }
    },
    'leg press': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_HIP, mp.solutions.pose.PoseLandmark.LEFT_KNEE, mp.solutions.pose.PoseLandmark.LEFT_ANKLE],
        'min_rep_angle': 90, 'max_rep_angle': 170, 'min_movement_diff': 60, 'target_rom': 80,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 2.0},
        'critical_angles': {'danger_zone': (75, 95)},
        'rep_threshold': 0.75, 'smoothing_factor': 5,
        'feedback': {
            'excellent': "Excellent leg press! Perfect 90° depth with controlled extension.",
            'good': "Good form. Maintain 90° knee angle at bottom.", 'average': "Average form. Focus on reaching 90° knee flexion.",
            'poor': "Poor form. Not reaching proper depth or control issues.", 'very_poor': "Dangerous form! Risk of knee injury - check depth and alignment.",
            'down': "Lower with control to 90° knee angle, feel the stretch.", 'up': "Drive through heels to full extension, don't lock knees.",
            'incomplete': "Incomplete rep. Must reach 90° knee flexion.",
            'form_violation': "Knee tracking or depth violation detected!"
        }
    },
    'seated dip machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 60, 'max_rep_angle': 165, 'min_movement_diff': 65, 'target_rom': 105,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 1.8},
        'critical_angles': {'danger_zone': (45, 65)},
        'rep_threshold': 0.75, 'smoothing_factor': 3,
        'feedback': {
            'excellent': "Outstanding dip form! Perfect tricep activation.",
            'good': "Good form. Keep elbows close to body.", 'average': "Average form. Focus on full range.",
            'poor': "Poor form. Control the descent.", 'very_poor': "Dangerous form! Risk of shoulder injury.",
            'down': "Lower with control, feel tricep stretch.", 'up': "Press up with power, full extension.",
            'incomplete': "Incomplete rep. Achieve full range.",
            'form_violation': "Elbow position needs correction!"
        }
    },
    'shoulder press machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_SHOULDER, mp.solutions.pose.PoseLandmark.LEFT_ELBOW, mp.solutions.pose.PoseLandmark.LEFT_WRIST],
        'min_rep_angle': 80, 'max_rep_angle': 165, 'min_movement_diff': 50, 'target_rom': 85,
        'ideal_tempo': {'eccentric': 2.0, 'concentric': 1.5},
        'critical_angles': {'danger_zone': (60, 85)},
        'rep_threshold': 0.7, 'smoothing_factor': 3,
        'feedback': {
            'excellent': "Perfect shoulder press form! Full range with control.",
            'good': "Good form. Minor adjustments needed.", 'average': "Decent form. Focus on full range of motion.",
            'poor': "Form needs improvement. Check your posture.", 'very_poor': "Stop! Risk of injury. Reset your form.",
            'down': "Lower with control, don't drop the weight.", 'up': "Press up with power, full extension.",
            'incomplete': "Incomplete rep. Ensure full range of motion.",
            'form_violation': "Form violation detected!"
        }
    },
    'smith machine': {
        'keypoints': [mp.solutions.pose.PoseLandmark.LEFT_HIP, mp.solutions.pose.PoseLandmark.LEFT_KNEE, mp.solutions.pose.PoseLandmark.LEFT_ANKLE],
        'min_rep_angle': 80, 'max_rep_angle': 170, 'min_movement_diff': 60, 'target_rom': 90,
        'ideal_tempo': {'eccentric': 2.5, 'concentric': 2.0},
        'critical_angles': {'danger_zone': (65, 85)},
        'rep_threshold': 0.75, 'smoothing_factor': 4,
        'feedback': {
            'excellent': "Excellent Smith machine squat! Perfect depth and form.",
            'good': "Good form. Maintain straight bar path.", 'average': "Average form. Focus on depth and control.",
            'poor': "Poor form. Check knee tracking and depth.", 'very_poor': "Dangerous form! Risk of knee/back injury.",
            'down': "Descend with control, hips back first.", 'up': "Drive through heels, full extension.",
            'incomplete': "Not deep enough. Aim for parallel or below.",
            'form_violation': "Knee tracking or depth issue detected!"
        }
    }
}

def calculate_iou(boxA: List[int], boxB: List[int]) -> float:
    try:
        xA = max(boxA[0], boxB[0]); yA = max(boxA[1], boxB[1])
        xB = min(boxA[2], boxB[2]); yB = min(boxA[3], boxB[3])
        if xB <= xA or yB <= yA: return 0.0
        inter_area = (xB - xA) * (yB - yA)
        boxA_area = (boxA[2] - boxA[0]) * (boxA[3] - boxA[1])
        boxB_area = (boxB[2] - boxB[0]) * (boxB[3] - boxB[1])
        union_area = boxA_area + boxB_area - inter_area
        return inter_area / union_area if union_area > 0 else 0.0
    except: return 0.0

def calculate_angle(a: Tuple[int, int], b: Tuple[int, int], c: Tuple[int, int]) -> float:
    try:
        a, b, c = np.array(a), np.array(b), np.array(c)
        ba = a - b; bc = c - b
        cos_angle = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc))
        cos_angle = np.clip(cos_angle, -1.0, 1.0)
        return np.degrees(np.arccos(cos_angle))
    except: return 90.0

def get_landmark_position(landmarks, landmark_id: int, frame_width: int, frame_height: int) -> Tuple[int, int, float]:
    try:
        lm = landmarks.landmark[landmark_id]
        x = max(0, min(int(lm.x * frame_width), frame_width - 1))
        y = max(0, min(int(lm.y * frame_height), frame_height - 1))
        return (x, y, lm.visibility)
    except: return (0, 0, 0.0)

def detect_form_violations(angle: float, rules: Dict, stage: str) -> List[str]:
    violations = []
    if angle < rules['min_rep_angle'] - 10: violations.append("Excessive range")
    if angle > rules['max_rep_angle'] + 10: violations.append("Insufficient range")
    
    critical_angles = rules.get('critical_angles', {})
    if 'danger_zone' in critical_angles:
        danger_min, danger_max = critical_angles['danger_zone']
        if danger_min <= angle <= danger_max:
            violations.append(rules['feedback'].get('form_violation', "Risk: Dangerous angle!"))
    return violations

def analyze_exercise_form_advanced(exercise_state: EnhancedExerciseState, landmarks, frame_shape) -> Optional[Tuple]:
    exercise_name = exercise_state.exercise_name
    rules = ENHANCED_EXERCISE_RULES.get(exercise_name)

    if not rules:
        exercise_state.feedback = "Exercise not recognized for analysis."
        return None

    h, w, _ = frame_shape
    keypoints_ids = rules['keypoints']
    
    p1_data = get_landmark_position(landmarks, keypoints_ids[0].value, w, h)
    p2_data = get_landmark_position(landmarks, keypoints_ids[1].value, w, h)
    p3_data = get_landmark_position(landmarks, keypoints_ids[2].value, w, h)
    
    if not all(p[2] > MIN_VISIBILITY_THRESHOLD for p in [p1_data, p2_data, p3_data]):
        exercise_state.feedback = "⚠️ Adjust position for better pose detection."
        return None
    
    current_angle = calculate_angle(p1_data[:2], p2_data[:2], p3_data[:2])
    
    smoothing_factor = rules.get('smoothing_factor', 3)
    smoothed_angle = exercise_state.update_angle_data(current_angle, smoothing_factor)
    
    exercise_state.min_angle_in_rep = min(exercise_state.min_angle_in_rep, smoothed_angle)
    exercise_state.max_angle_in_rep = max(exercise_state.max_angle_in_rep, smoothed_angle)
    violations = detect_form_violations(smoothed_angle, rules, exercise_state.stage)
    exercise_state.form_violations.extend(violations)
    posture_metrics = exercise_state.evaluate_posture_quality(rules)
    
    exercise_state.form_score_history.append(posture_metrics.overall_score)
    
    current_time = time.time()
    
    if smoothed_angle < rules['min_rep_angle']:
        if exercise_state.stage != "down":
            exercise_state.stage = "down"
            exercise_state.concentric_duration = current_time - exercise_state.phase_start_time
            exercise_state.phase_start_time = current_time
            exercise_state.feedback = rules['feedback'].get('down', "Lower with control.")
            
    elif smoothed_angle > rules['max_rep_angle'] and exercise_state.stage == "down":
        exercise_state.stage = "up"
        exercise_state.eccentric_duration = current_time - exercise_state.phase_start_time
        exercise_state.phase_start_time = current_time
        
        range_of_motion = exercise_state.max_angle_in_rep - exercise_state.min_angle_in_rep
        
        rep_threshold = rules.get('rep_threshold', 0.75)
        if range_of_motion >= (rules['target_rom'] * rep_threshold):
            exercise_state.reps += 1
            quality_feedback = rules['feedback'].get(posture_metrics.quality.value.lower(), rules['feedback']['average'])
            exercise_state.feedback = f"🎉 Rep {exercise_state.reps}! {quality_feedback}"
        else:
            exercise_state.feedback = rules['feedback'].get('incomplete', "Incomplete Rep.")
        
        exercise_state.reset_rep_cycle()

    return (p1_data[:2], p2_data[:2], p3_data[:2], smoothed_angle, posture_metrics)

def draw_advanced_overlay(frame, person, state: EnhancedExerciseState, analysis_result):
    p_box = person['box']
    color_map = {
        PostureQuality.EXCELLENT: (0, 255, 0), PostureQuality.GOOD: (50, 205, 50),
        PostureQuality.AVERAGE: (255, 191, 0), PostureQuality.POOR: (255, 140, 0),
        PostureQuality.VERY_POOR: (255, 0, 0)
    }
    quality_color = color_map.get(state.posture_metrics.quality, (255, 255, 255))
    cv2.rectangle(frame, (p_box[0], p_box[1]), (p_box[2], p_box[3]), quality_color, 2)
    
    if analysis_result and len(analysis_result) == 5:
        p1, p2, p3, angle, _ = analysis_result
        cv2.line(frame, p1, p2, quality_color, 3); cv2.line(frame, p2, p3, quality_color, 3)
        cv2.circle(frame, p2, 8, quality_color, -1)

    info_y = p_box[1] - 10; font_scale = 0.5; thickness = 1
    quality_emoji_map = {
        PostureQuality.EXCELLENT: "✅", PostureQuality.GOOD: "👍",
        PostureQuality.AVERAGE: "👌", PostureQuality.POOR: "⚠️",
        PostureQuality.VERY_POOR: "❌"
    }
    quality_emoji = quality_emoji_map.get(state.posture_metrics.quality, "")
    
    score_text = f"📊 Score: {state.posture_metrics.overall_score:.0f}/100"
    cv2.putText(frame, score_text, (p_box[0], info_y - 35), cv2.FONT_HERSHEY_SIMPLEX, font_scale, quality_color, thickness+1)
    
    exercise_text = f"🏋️ {state.exercise_name} | Reps: {state.reps}"
    cv2.putText(frame, exercise_text, (p_box[0], info_y - 18), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (0, 255, 255), thickness+1)
    
    info_text = f"ID: {state.person_id} | {state.posture_metrics.quality.value} {quality_emoji}"
    cv2.putText(frame, info_text, (p_box[0], info_y), cv2.FONT_HERSHEY_SIMPLEX, font_scale, (255, 255, 255), thickness)
    
    feedback = state.feedback
    if len(feedback) > 50: feedback = feedback[:50] + "..."
    cv2.putText(frame, feedback, (p_box[0], p_box[3] + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 255), 1)

def save_stats_to_excel(session_stats: Dict, person_exercise_states: Dict[int, EnhancedExerciseState], filename: str):
    filepath = os.path.join(tempfile.gettempdir(), filename)
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames: wb.remove(wb['Sheet'])

    ws_summary = wb.create_sheet("Session Summary", 0)
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Report Generated", time.strftime('%Y-%m-%d %H:%M:%S')])
    ws_summary.append(["Processing Time (seconds)", f"{session_stats['processing_time']:.2f}"])
    ws_summary.append(["Total Reps", session_stats['total_reps']])
    ws_summary.append(["Exercises Performed", ", ".join(session_stats['exercises_performed'])])
    ws_summary.append(["Average Form Score", f"{session_stats['average_form_score']:.1f}"])
    for col in ws_summary.columns: ws_summary.column_dimensions[col[0].column_letter].width = 30
    for row in ws_summary[1]: row.font = openpyxl.styles.Font(bold=True)
    
    ws_details = wb.create_sheet("Detailed Person Stats", 1)
    headers = ["Person ID", "Exercise Name", "Total Reps", "Overall Quality", "Overall Score", "Angle Consistency", "Range of Motion", "Form Stability", "Timing Consistency", "Tempo Score", "Avg Rep Time (s)"]
    ws_details.append(headers)
    for person_id, state in person_exercise_states.items():
        if state.reps > 0:
            avg_rep_time = np.mean(list(state.rep_durations)) if state.rep_durations else 0.0
            row_data = [
                state.person_id, state.exercise_name, state.reps, state.posture_metrics.quality.value,
                f"{state.posture_metrics.overall_score:.1f}", f"{state.posture_metrics.angle_consistency:.1f}",
                f"{state.posture_metrics.range_of_motion:.1f}", f"{state.posture_metrics.form_stability:.1f}",
                f"{state.posture_metrics.timing_consistency:.1f}", f"{state.posture_metrics.tempo_score:.1f}",
                f"{avg_rep_time:.2f}"
            ]
            ws_details.append(row_data)
    for col in ws_details.columns: ws_details.column_dimensions[col[0].column_letter].width = 22
    for cell in ws_details[1]: cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF"); cell.fill = openpyxl.styles.PatternFill("solid", fgColor="4F81BD")

    try:
        wb.save(filepath)
        return filepath
    except Exception as e:
        st.error(f"Error saving excel file: {e}")
        return None

def generate_form_score_graph(person_exercise_states: Dict[int, EnhancedExerciseState]):
    active_users = {pid: state for pid, state in person_exercise_states.items() if state.reps > 0 and state.form_score_history}
    if not active_users:
        return None

    plt.style.use('seaborn-v0_8-darkgrid')
    fig, ax = plt.subplots(figsize=(12, 7))

    for person_id, state in active_users.items():
        scores = state.form_score_history[::10]
        frames = range(0, len(state.form_score_history), 10)
        ax.plot(frames, scores, marker='o', linestyle='-', markersize=4, label=f'Person {person_id} ({state.exercise_name})')

    ax.set_ylabel("Overall Form Score (0-100)", fontsize=12)
    ax.set_xlabel("Frames (timeline of workout)", fontsize=12)
    ax.set_title("Form Score Over Time", fontsize=16, pad=20)
    ax.set_ylim(0, 105)
    ax.legend()
    plt.tight_layout()
    return fig

def run_streamlit_inference_machine(person_conf, equipment_conf, iou_threshold):
    person_model, equipment_model = load_models()
    equipment_class_names = load_equipment_classes()
    if not all([person_model, equipment_model, equipment_class_names]): st.stop()

    cap = cv2.VideoCapture(st.session_state.video_source)
    if not cap.isOpened():
        st.error("Error opening video source."); st.session_state.run_analysis = False; st.rerun(); return

    if st.session_state.output_video_path is None:
        temp_dir = tempfile.gettempdir()
        unique_filename = f"processed_video_machine_{int(time.time())}.mp4"
        st.session_state.output_video_path = os.path.join(temp_dir, unique_filename)

    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(st.session_state.output_video_path, fourcc, cap.get(cv2.CAP_PROP_FPS), (int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)), int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))))

    image_placeholder = st.empty()
    st.divider()
    st.subheader("📊 Live Metrics")
    col1, col2, col3 = st.columns(3)
    metric1_placeholder = col1.empty()
    metric2_placeholder = col2.empty()
    metric3_placeholder = col3.empty()
    
    machine_detections_cache = []
    person_detections_cache = []
    person_assignments_cache = {}
    pose_landmarks_cache = None
    frame_idx = 0
    
    with mp.solutions.pose.Pose(min_detection_confidence=0.7, min_tracking_confidence=0.7, model_complexity=1) as pose:
        session_start_time = time.time()
        
        while st.session_state.run_analysis and cap.isOpened():
            ret, frame = cap.read()
            if not ret: break
            frame_idx += 1

            if frame_idx % 2 == 0:
                equipment_results = equipment_model(frame, verbose=False, conf=equipment_conf)
                machine_detections_cache = [{'box': list(map(int, box.xyxy[0])), 'name': equipment_class_names[int(box.cls[0])], 'occupied': False} for r in equipment_results for box in r.boxes]

                person_results = person_model.track(frame, persist=True, verbose=False, conf=person_conf, classes=[0])
                person_detections_cache = []
                if person_results[0].boxes.id is not None:
                    person_detections_cache = [{'box': b, 'id': i} for b, i in zip(person_results[0].boxes.xyxy.cpu().numpy().astype(int), person_results[0].boxes.id.cpu().numpy().astype(int))]

                person_assignments_cache.clear()
                for machine in machine_detections_cache:
                    best_iou, assigned_person_id = 0, None
                    for person in person_detections_cache:
                        iou = calculate_iou(machine['box'], person['box'])
                        if iou > iou_threshold and iou > best_iou: best_iou, assigned_person_id = iou, person['id']
                    if assigned_person_id: machine['occupied'], person_assignments_cache[assigned_person_id] = True, machine['name']
                
                pose_results = pose.process(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))
                pose_landmarks_cache = pose_results.pose_landmarks if pose_results.pose_landmarks else None
            
            display_frame = frame.copy()
            
            current_total_reps = 0
            current_avg_form_score = 0
            active_exercises = set()

            for person in person_detections_cache:
                person_id = person['id']
                if person_id not in st.session_state.person_exercise_states:
                    st.session_state.person_exercise_states[person_id] = EnhancedExerciseState(person_id)

                state = st.session_state.person_exercise_states[person_id]
                assigned_machine_name = person_assignments_cache.get(person_id, "Free Movement")
                state.exercise_name = assigned_machine_name  

                if pose_landmarks_cache:
                    analysis_result = analyze_exercise_form_advanced(state, pose_landmarks_cache, frame.shape)
                    if analysis_result:
                        mp.solutions.drawing_utils.draw_landmarks(
                            display_frame,
                            pose_landmarks_cache,
                            mp.solutions.pose.POSE_CONNECTIONS,
                            landmark_drawing_spec=mp.solutions.drawing_utils.DrawingSpec(color=(0, 255, 0), thickness=2, circle_radius=2),
                            connection_drawing_spec=mp.solutions.drawing_utils.DrawingSpec(color=(0, 0, 255), thickness=2, circle_radius=2)
                        )
                        draw_advanced_overlay(display_frame, person, state, analysis_result)
                
                if state.reps > 0:
                    current_total_reps += state.reps
                    current_avg_form_score += state.posture_metrics.overall_score
                    active_exercises.add(state.exercise_name)

            if len(person_detections_cache) > 0 and current_total_reps > 0:
                active_people_count = len([s for s in st.session_state.person_exercise_states.values() if s.reps > 0])
                if active_people_count > 0:
                    current_avg_form_score /= active_people_count

            metric1_placeholder.metric(label="Total Reps", value=current_total_reps)
            metric2_placeholder.metric(label="Avg. Form Score", value=f"{current_avg_form_score:.1f}")
            metric3_placeholder.metric(label="Exercises", value=len(active_exercises))

            for machine in machine_detections_cache:
                color = (0, 255, 0) if machine['occupied'] else (0, 165, 255) 
                text = f"{machine['name']} - {'Occupied' if machine['occupied'] else 'Free'}"
                cv2.rectangle(display_frame, (machine['box'][0], machine['box'][1]), (machine['box'][2], machine['box'][3]), color, 2)
                cv2.putText(display_frame, text, (machine['box'][0], machine['box'][1] - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

            image_placeholder.image(display_frame, channels="BGR", use_container_width=True)
            out.write(display_frame)
        
        cap.release()
        out.release()
        total_reps = sum(state.reps for state in st.session_state.person_exercise_states.values())
        avg_form_score = np.mean([state.posture_metrics.overall_score for state in st.session_state.person_exercise_states.values() if state.reps > 0]) if total_reps > 0 else 0
        exercises_performed = list(set(state.exercise_name for state in st.session_state.person_exercise_states.values() if state.reps > 0))
        
        st.session_state.session_stats = {
            "processing_time": time.time() - session_start_time,
            "total_reps": total_reps,
            "average_form_score": avg_form_score,
            "exercises_performed": exercises_performed
        }
        st.session_state.excel_report_path = save_stats_to_excel(st.session_state.session_stats, st.session_state.person_exercise_states, "gym_analysis_report.xlsx")
        st.session_state.graph_fig = generate_form_score_graph(st.session_state.person_exercise_states)
        
        st.session_state.analysis_complete = True
        st.session_state.run_analysis = False 
        st.rerun()
EXTENDED, FLEXED = 0, 1

class RepetitionFSM:
    def __init__(self, flex_thresh, ext_thresh, margin=5):
        self.flex_thresh = flex_thresh
        self.ext_thresh = ext_thresh
        self.margin = margin
        self.state = EXTENDED
        self.count = 0

    def update(self, angle):
        if self.state == EXTENDED and angle < (self.flex_thresh - self.margin):
            self.state = FLEXED
        elif self.state == FLEXED and angle > (self.ext_thresh + self.margin):
            self.state = EXTENDED
            self.count += 1
        return self.count

def compute_angle(a, b, c):
    ba = np.array([a.x - b.x, a.y - b.y])
    bc = np.array([c.x - b.x, c.y - b.y])
    cosang = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-6)
    return np.degrees(np.arccos(np.clip(cosang, -1, 1)))

def get_landmark_visibility(landmark):
    return hasattr(landmark, 'visibility') and landmark.visibility > 0.5

def generate_non_machine_graphs(frame_data: List[Dict]):
    if not frame_data:
        return None, None

    df = pd.DataFrame(frame_data)
    plt.style.use('seaborn-v0_8-darkgrid')
    rep_durations_data = []
    for exercise_type in df['Exercise'].unique():
        exercise_df = df[df['Exercise'] == exercise_type]
        rep_completion_df = exercise_df[exercise_df['Reps'].diff() == 1]
        completion_times = rep_completion_df['Timestamp'].tolist()
        if not completion_times:
            continue
        
        start_frames = exercise_df[exercise_df['Reps'] == 0]
        last_time = start_frames.iloc[-1]['Timestamp'] if not start_frames.empty else exercise_df.iloc[0]['Timestamp']

        for i, t in enumerate(completion_times):
            duration = t - last_time
            rep_durations_data.append({'Exercise': exercise_type.capitalize(), 'Rep': i + 1, 'Duration (s)': duration})
            last_time = t

    fig1, ax1 = plt.subplots(figsize=(10, 5))
    if rep_durations_data:
        duration_df = pd.DataFrame(rep_durations_data)
        sns.barplot(data=duration_df, x='Rep', y='Duration (s)', hue='Exercise', ax=ax1)
        ax1.set_title("Repetition Duration Analysis")
        ax1.set_xlabel("Repetition Number")
        ax1.set_ylabel("Processing Duration (seconds)")
        ax1.legend(title='Exercise')
        ax1.grid(True, axis='y', linestyle='--')
    else:
        ax1.text(0.5, 0.5, 'Not enough data for Rep Duration graph.', ha='center', va='center')
    fig2, ax2 = plt.subplots(figsize=(12, 6))
    if 'Form_Score' in df.columns and not df.empty:
        df['Time_Bin'] = pd.to_numeric(df['Timestamp'], errors='coerce').fillna(0).astype(int) // 2 * 2
        grouped_data = df.groupby(['Time_Bin', 'Exercise'])['Form_Score'].mean().reset_index()

        for exercise in grouped_data['Exercise'].unique():
            exercise_data = grouped_data[grouped_data['Exercise'] == exercise]
            ax2.plot(exercise_data['Time_Bin'], exercise_data['Form_Score'], marker='o', linestyle='-', label=f'{exercise.capitalize()} Form')

        ax2.set_xlabel("Time (seconds into workout)")
        ax2.set_ylabel("Average Form Score (0-100)")
        ax2.set_title("Form Score Over Time (Sampled every 2 seconds)")
        ax2.set_ylim(0, 105)
        ax2.legend()
        ax2.grid(True, linestyle='--')
    else:
        ax2.text(0.5, 0.5, 'No Form Score data available.', ha='center', va='center')
    fig1.tight_layout()
    fig2.tight_layout()

    return fig1, fig2
class ExcelReporter_NonMachine:
    def __init__(self):
        self.session_data = []
        self.frame_data = []
        self.start_time = None
        self.end_time = None

    def start_session(self):
        self.start_time = time.time()

    def end_session(self):
        self.end_time = time.time()

    def add_frame_data(self, frame_number, exercise_type, reps, angle, feedback, view, form_score):
        timestamp = time.time() - self.start_time if self.start_time else 0
        self.frame_data.append({
            'Frame': frame_number,
            'Timestamp': timestamp,
            'Exercise': exercise_type,
            'Reps': reps,
            'Angle': angle,
            'Feedback': feedback,
            'View': view,
            'Form_Score': form_score
        })

    def add_session_summary(self, video_path, video_length, processing_time, pushup_count, squat_count):
        self.session_data.append({
            'Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'Video_File': os.path.basename(video_path),
            'Video_Length_sec': video_length,
            'Processing_Time_sec': processing_time,
            'Total_Pushups': pushup_count,
            'Total_Squats': squat_count,
            'Pushups_per_min': (pushup_count / video_length * 60) if video_length > 0 else 0,
            'Squats_per_min': (squat_count / video_length * 60) if video_length > 0 else 0,
            'Processing_Speed': (video_length / processing_time) if processing_time > 0 else 0
        })

    def export_to_excel(self, filename='exercise_analysis.xlsx'):
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            if self.session_data:
                session_df = pd.DataFrame(self.session_data)
                session_df.to_excel(writer, sheet_name='Session_Summary', index=False)
            if self.frame_data:
                frame_df = pd.DataFrame(self.frame_data)
                frame_df.to_excel(writer, sheet_name='Frame_Data', index=False)
                rep_progression = frame_df.groupby('Exercise')['Reps'].max().reset_index()
                rep_progression.to_excel(writer, sheet_name='Rep_Summary', index=False)
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.load_workbook(filename)
            if 'Session_Summary' in wb.sheetnames:
                ws = wb['Session_Summary']
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            if 'Frame_Data' in wb.sheetnames:
                ws = wb['Frame_Data']
                header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            wb.save(filename)
        except Exception:
            pass
        return filename

class MultiViewExerciseTracker:
    def __init__(self, pushup_params, squat_params):
        self.pushup_fsm = RepetitionFSM(**pushup_params)
        self.squat_fsm = RepetitionFSM(**squat_params)
        self.mode = None
        self.view = None
        self.excel_reporter = ExcelReporter_NonMachine()
        self.frame_count = 0

    def detect_view_and_mode(self, lm):
        y_wrists = (lm[mp.solutions.pose.PoseLandmark.LEFT_WRIST].y +
                   lm[mp.solutions.pose.PoseLandmark.RIGHT_WRIST].y) / 2
        y_nose = lm[mp.solutions.pose.PoseLandmark.NOSE].y
        y_lower = (lm[mp.solutions.pose.PoseLandmark.LEFT_HIP].y +
                   lm[mp.solutions.pose.PoseLandmark.LEFT_KNEE].y +
                   lm[mp.solutions.pose.PoseLandmark.RIGHT_HIP].y +
                   lm[mp.solutions.pose.PoseLandmark.RIGHT_KNEE].y) / 4
        mode = 'squat' if (y_wrists < y_lower and y_nose < y_lower) else 'pushup'
        left_sh = lm[mp.solutions.pose.PoseLandmark.LEFT_SHOULDER]
        right_sh = lm[mp.solutions.pose.PoseLandmark.RIGHT_SHOULDER]
        shoulder_width = abs(left_sh.x - right_sh.x)
        view = 'side' if shoulder_width < 0.1 else 'front'
        return mode, view

    def get_feedback(self, angle, mode):
        if mode == 'pushup':
            if angle < (self.pushup_fsm.flex_thresh - self.pushup_fsm.margin):
                return 'Go lower'
            elif angle > (self.pushup_fsm.ext_thresh + self.pushup_fsm.margin):
                return 'Extend more'
            else:
                return 'Good push-up form'
        else:
            if angle < (self.squat_fsm.flex_thresh - self.squat_fsm.margin):
                return 'Go deeper'
            elif angle > (self.squat_fsm.ext_thresh + self.squat_fsm.margin):
                return 'Stand fully'
            else:
                return 'Good squat form'

    def process_video(self, input_path, output_path, detection_conf, tracking_conf,
                      feedback_container, image_container, metrics_containers):
        cap = cv2.VideoCapture(input_path)
        width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = cap.get(cv2.CAP_PROP_FPS) or 30.0
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        video_length = total_frames / fps if fps > 0 else 0

        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        writer = cv2.VideoWriter(output_path, fourcc, fps, (width, height))

        pose = mp.solutions.pose.Pose(static_image_mode=False,
                            model_complexity=1,
                            enable_segmentation=False,
                            min_detection_confidence=detection_conf,
                            min_tracking_confidence=tracking_conf)

        self.excel_reporter.start_session()
        start_time = time.time()

        progress_bar = st.progress(0)
        frame_idx = 0

        while True:
            ret, frame = cap.read()
            if not ret:
                break
            frame_idx += 1

            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            res = pose.process(rgb)
            output_text = 'No person detected'

            if res.pose_landmarks:
                lm = res.pose_landmarks.landmark
                mp.solutions.drawing_utils.draw_landmarks(
                    frame,
                    res.pose_landmarks,
                    mp.solutions.pose.POSE_CONNECTIONS,
                    mp.solutions.drawing_utils.DrawingSpec(color=(0,255,0), thickness=2, circle_radius=2),
                    mp.solutions.drawing_utils.DrawingSpec(color=(0,0,255), thickness=2)
                )
                new_mode, new_view = self.detect_view_and_mode(lm)
                if new_mode != self.mode:
                    self.mode = new_mode
                    self.view = new_view
                    
                    self.pushup_fsm.count = 0; self.pushup_fsm.state = EXTENDED
                    self.squat_fsm.count = 0; self.squat_fsm.state = EXTENDED

                if self.mode == 'pushup':
                    angle = compute_angle(
                        lm[mp.solutions.pose.PoseLandmark.LEFT_SHOULDER],
                        lm[mp.solutions.pose.PoseLandmark.LEFT_ELBOW],
                        lm[mp.solutions.pose.PoseLandmark.LEFT_WRIST]
                    )
                    count = self.pushup_fsm.update(angle)
                else:
                    angle = compute_angle(
                        lm[mp.solutions.pose.PoseLandmark.LEFT_HIP],
                        lm[mp.solutions.pose.PoseLandmark.LEFT_KNEE],
                        lm[mp.solutions.pose.PoseLandmark.LEFT_ANKLE]
                    )
                    count = self.squat_fsm.update(angle)

                feedback = self.get_feedback(angle, self.mode)
                
                if 'Good' in feedback:
                    form_score = 90
                elif 'lower' in feedback or 'deeper' in feedback:
                    form_score = 60
                elif 'Extend' in feedback or 'fully' in feedback:
                    form_score = 70
                else:
                    form_score = 50

                output_text = f'{self.mode.capitalize()} Count: {count}, Angle: {angle:.2f}, Feedback: {feedback}'
                feedback_container.write(output_text)

                metrics_containers[0].metric(label=f"Total {self.mode.capitalize()}s", value=count)
                
                self.excel_reporter.add_frame_data(frame_idx, self.mode, count, angle, feedback, self.view, form_score)

            image_container.image(frame, channels="BGR", use_container_width=True)
            writer.write(frame)
            progress_bar.progress(min(100, int((frame_idx / total_frames) * 100)))

        cap.release()
        writer.release()
        pose.close()

        end_time = time.time()
        processing_time = end_time - start_time
        
        total_pushups = self.pushup_fsm.count
        total_squats = self.squat_fsm.count
        
        self.excel_reporter.end_session()
        excel_path = self.excel_reporter.export_to_excel(filename=f"non_machine_analysis_report_{int(time.time())}.xlsx")
        
        rep_duration_graph, form_score_graph = generate_non_machine_graphs(self.excel_reporter.frame_data)

        return output_path, excel_path, total_pushups, total_squats, video_length, processing_time, rep_duration_graph, form_score_graph

def render_non_machine_analysis_page():
    st.header("🤸 Non-Machine Exercise Analysis")
    st.write("Upload a video of your non-machine exercises (e.g., push-ups, squats) for real-time form analysis.")

    detection_conf = st.slider("Detection Confidence", 0.1, 1.0, 0.5, 0.05)
    tracking_conf = st.slider("Tracking Confidence", 0.1, 1.0, 0.5, 0.05)

    pushup_flex_thresh = st.slider("Push-up: Flexed Angle Threshold", 0, 180, 50)
    pushup_ext_thresh = st.slider("Push-up: Extended Angle Threshold", 0, 180, 160)
    squat_flex_thresh = st.slider("Squat: Flexed Angle Threshold", 0, 180, 80)
    squat_ext_thresh = st.slider("Squat: Extended Angle Threshold", 0, 180, 160)

    pushup_params = {'flex_thresh': pushup_flex_thresh, 'ext_thresh': pushup_ext_thresh}
    squat_params = {'flex_thresh': squat_flex_thresh, 'ext_thresh': squat_ext_thresh}
    
    uploaded_file = st.file_uploader("Upload a video", type=["mp4", "mov", "avi", "mkv"])

    if uploaded_file is not None:
        temp_video_path = os.path.join(tempfile.gettempdir(), uploaded_file.name)
        with open(temp_video_path, "wb") as f:
            f.write(uploaded_file.read())
        
        st.video(temp_video_path, format="video/mp4", start_time=0)
        
        if st.button("Start Analysis", key="start_non_machine_analysis"):
            st.session_state.run_non_machine_analysis = True
            st.session_state.non_machine_video_source = temp_video_path
            st.session_state.non_machine_output_path = os.path.join(tempfile.gettempdir(), f"processed_non_machine_{int(time.time())}.mp4")
            st.session_state.non_machine_analysis_complete = False
            st.session_state.non_machine_report = None
            st.session_state.pushup_count = 0
            st.session_state.squat_count = 0
            st.rerun()

    if st.session_state.get('run_non_machine_analysis', False) and st.session_state.get('non_machine_video_source'):
        feedback_placeholder = st.empty()
        video_placeholder = st.empty()
        col1, col2 = st.columns(2)
        metrics_containers = [col1.empty(), col2.empty()]
        
        tracker = MultiViewExerciseTracker(pushup_params, squat_params)
        
        processed_video, excel_report, total_pushups, total_squats, video_length, processing_time, rep_duration_fig, form_score_fig = tracker.process_video(
            st.session_state.non_machine_video_source,
            st.session_state.non_machine_output_path,
            detection_conf,
            tracking_conf,
            feedback_placeholder,
            video_placeholder,
            metrics_containers
        )
        
        st.session_state.non_machine_analysis_complete = True
        st.session_state.non_machine_report = {
            "processed_video": processed_video,
            "excel_report": excel_report,
            "total_pushups": total_pushups,
            "total_squats": total_squats,
            "video_length": video_length,
            "processing_time": processing_time,
            "rep_duration_graph": rep_duration_fig,
            "form_score_graph": form_score_fig
        }
        st.session_state.run_non_machine_analysis = False
        st.rerun()

    if st.session_state.get('non_machine_analysis_complete', False) and st.session_state.get('non_machine_report'):
        report = st.session_state.non_machine_report
        st.success("✅ Analysis Complete!")
        st.video(report['processed_video'])
        st.download_button(
            label="Download Processed Video",
            data=open(report['processed_video'], "rb").read(),
            file_name=os.path.basename(report['processed_video']),
            mime="video/mp4"
        )
        if report['excel_report']:
            st.download_button(
                label="Download Analysis Report (Excel)",
                data=open(report['excel_report'], "rb").read(),
                file_name=os.path.basename(report['excel_report']),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.subheader("Summary Report")
        st.write(f"**Total Push-ups:** {report['total_pushups']}")
        st.write(f"**Total Squats:** {report['total_squats']}")
        st.write(f"**Video Length:** {report['video_length']:.2f} seconds")
        st.write(f"**Processing Time:** {report['processing_time']:.2f} seconds")

        st.write("---")
        st.subheader("📊 Graphical Analysis")

        rep_graph = report.get('rep_duration_graph')
        form_graph = report.get('form_score_graph')
        
        if rep_graph and form_graph:
            col_g1, col_g2 = st.columns(2)
            with col_g1:
                st.markdown("#### Repetition Duration")
                st.pyplot(rep_graph)
            with col_g2:
                st.markdown("#### Form Score Over Time")
                st.pyplot(form_graph)
        elif rep_graph:
            st.markdown("#### Repetition Duration")
            st.pyplot(rep_graph)
        elif form_graph:
            st.markdown("#### Form Score Over Time")
            st.pyplot(form_graph)
        else:
            st.info("No graphical data available for this session.")

class ExerciseTrackerEfficiency:
    def __init__(self):
        self.person_machine_data = {}
        self.exercise_records = []
        self.pose_history = {}
        self.rep_state = {}
        self.rep_count = {}
        self.feedback_data = {}
        self.machine_efficiency = {}  

    def update_person_machine(self, person_id, machine_name, efficiency_status=None): 
        if person_id not in self.person_machine_data:
            self.person_machine_data[person_id] = {}
        self.person_machine_data[person_id][machine_name] = True

        if efficiency_status:
            if person_id not in self.machine_efficiency:
                self.machine_efficiency[person_id] = {}
            self.machine_efficiency[person_id][machine_name] = efficiency_status

    def record_exercise_session(self, person_id, machine_name): 
        efficiency = self.machine_efficiency.get(person_id, {}).get(machine_name, "Unknown")
        self.exercise_records.append({
            "Person ID": person_id,
            "Machine": machine_name,
            "Efficiency": efficiency
        })

    def save_to_excel(self, filename):
        for person_id, machines_data in list(self.person_machine_data.items()):
            for machine_name in list(machines_data.keys()):
                self.record_exercise_session(person_id, machine_name)

        if self.exercise_records:
            df = pd.DataFrame(self.exercise_records)
            df.to_excel(filename, index=False)
            print(f"Exercise data saved to {filename}")
        else:
            print("No exercise data to save.")

    def smooth_pose_landmarks(self, person_id, landmarks):
        if person_id not in self.pose_history:
            self.pose_history[person_id] = deque(maxlen=5)
        self.pose_history[person_id].append(landmarks)
        return self.pose_history[person_id][-1]

    def calculate_angle(self, a, b, c):
        a = np.array(a)
        b = np.array(b)
        c = np.array(c)
        ba = a - b
        bc = c - b
        cosine_angle = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-6)
        angle = np.arccos(np.clip(cosine_angle, -1.0, 1.0))
        return np.degrees(angle)

    def analyze_machine_exercise(self, landmarks, person_id, machine_name):
        left_shoulder = [landmarks.landmark[11].x, landmarks.landmark[11].y]
        left_elbow = [landmarks.landmark[13].x, landmarks.landmark[13].y]
        left_wrist = [landmarks.landmark[15].x, landmarks.landmark[15].y]
        angle = self.calculate_angle(left_shoulder, left_elbow, left_wrist)

        if person_id not in self.rep_state:
            self.rep_state[person_id] = "down"
            self.rep_count[person_id] = 0
            self.feedback_data[person_id] = []

        state = self.rep_state[person_id]
        reps = self.rep_count[person_id]
        feedback = []

        if angle < 50 and state == "down":
            self.rep_state[person_id] = "up"
        elif angle > 160 and state == "up":
            self.rep_state[person_id] = "down"
            reps += 1
            self.rep_count[person_id] = reps

        if angle > 170:
            feedback.append("Arm fully extended")
        elif angle < 40:
            feedback.append("Full contraction")
        else:
            feedback.append("Incomplete rep")

        self.feedback_data[person_id] = feedback

        return {
            "reps": reps,
            "angle": angle,
            "state": state,
            "feedback": feedback,
            "machine": machine_name,
            "active": True
        }

class DetectionStabilizerEfficiency:
    def __init__(self):
        self.trackers = {"machine": {}, "human": {}}

    def get_stable_detections(self, detections, dtype):
        for i, d in enumerate(detections):
            d["track_id"] = f"{dtype[0].upper()}_{i}"
        return detections

def calculate_occupancy_score(region, buffer, occupancy_threshold_val):
    gray = cv2.cvtColor(region, cv2.COLOR_BGR2GRAY)
    laplacian = cv2.Laplacian(gray, cv2.CV_64F)
    variance = laplacian.var()

    buffer.append(variance)
    if len(buffer) > 10:
        buffer.pop(0)

    mean_var = np.mean(buffer)

    if mean_var > 15:
        return "Occupied - Efficient"
    elif mean_var > occupancy_threshold_val: 
        return "Occupied - Inefficient"
    else:
        return "No Activity"

def assign_humans_to_machines_efficiency(humans, machines, iou_threshold):
    assignments = {}
    for human in humans:
        human_box = human["box"]
        hx_center = (human_box[0] + human_box[2]) / 2
        hy_center = (human_box[1] + human_box[3]) / 2
        best_match = None
        best_iou = 0.0

        for machine in machines:
            mx1, my1, mx2, my2 = machine["box"]
            if mx1 <= hx_center <= mx2 and my1 <= hy_center <= my2:
                iou = 1.0
            else:
                inter_x1 = max(human_box[0], mx1)
                inter_y1 = max(human_box[1], my1)
                inter_x2 = min(human_box[2], mx2)
                inter_y2 = min(human_box[3], my2)
                if inter_x2 > inter_x1 and inter_y2 > inter_y1:
                    intersection_area = (inter_x2 - inter_x1) * (inter_y2 - inter_y1)
                    human_area = (human_box[2] - human_box[0]) * (human_box[3] - human_box[1])
                    machine_area = (mx2 - mx1) * (my2 - my1)
                    union_area = human_area + machine_area - intersection_area
                    iou = intersection_area / union_area
                else:
                    iou = 0.0

            if iou > best_iou and iou > iou_threshold:
                best_iou = iou
                best_match = machine["class"]

        if best_match:
            assignments[human["track_id"]] = best_match

    return assignments

def draw_results_with_laplacian(image, machines, humans, analysis_data, assignments, laplacian_history, frame_idx, occupancy_threshold_val):
    annotated = image.copy()

    for machine in machines:
        x1, y1, x2, y2 = map(int, machine["box"])
        machine_crop = annotated[y1:y2, x1:x2]
        track_id = machine.get("track_id", "Unknown")

        if track_id not in laplacian_history:
            laplacian_history[track_id] = deque(maxlen=10)

        status = calculate_occupancy_score(machine_crop, laplacian_history[track_id], occupancy_threshold_val)
        color = (0, 255, 0) if "Efficient" in status else ((0, 165, 255) if "Inefficient" in status else (0, 0, 255))

        cv2.rectangle(annotated, (x1, y1), (x2, y2), color, 2)
        cv2.putText(annotated, f"{machine['class']} - {status}", (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

    for human in humans:
        x1, y1, x2, y2 = map(int, human["box"])
        track_id = human.get("track_id", "Unknown")
        analysis = analysis_data.get(track_id, {})
        feedback_text = ", ".join(analysis.get("feedback", []))
        reps = analysis.get("reps", 0)

        cv2.rectangle(annotated, (x1, y1), (x2, y2), (255, 0, 0), 2)
        cv2.putText(annotated, f"ID: {track_id} | Reps: {reps}", (x1, y1 - 25),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 0, 0), 2)
        cv2.putText(annotated, feedback_text, (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, (200, 200, 200), 1)

    return annotated, laplacian_history

def process_frame_efficiency(image, exercise_tracker_eff, stabilizer_eff, laplacian_history, frame_idx, conf_threshold_eff, occupancy_threshold_val):
    human_model = YOLO(PERSON_DETECTION_MODEL_PATH)
    gym_model = YOLO(GYM_EQUIPMENT_MODEL_PATH)

    machine_results = gym_model.predict(image, conf=conf_threshold_eff * 1.1, verbose=False)
    human_results = human_model.predict(image, conf=conf_threshold_eff, verbose=False)

    raw_machines, raw_humans = [], []

    if machine_results[0].boxes is not None:
        boxes = machine_results[0].boxes.xyxy.cpu().numpy()
        confidences = machine_results[0].boxes.conf.cpu().numpy()
        class_ids = machine_results[0].boxes.cls.cpu().numpy().astype(int)
        class_names = gym_model.names
        for box, conf, class_id in zip(boxes, confidences, class_ids):
            raw_machines.append({"box": box, "conf": conf, "class": class_names[class_id]})

    if human_results[0].boxes is not None:
        boxes = human_results[0].boxes.xyxy.cpu().numpy()
        confidences = human_results[0].boxes.conf.cpu().numpy()
        class_ids = human_results[0].boxes.cls.cpu().numpy().astype(int)
        for box, conf, class_id in zip(boxes, confidences, class_ids):
            if class_id == 0:
                raw_humans.append({"box": box, "conf": conf, "class": "person"})

    machines = stabilizer_eff.get_stable_detections(raw_machines, "machine")
    humans = stabilizer_eff.get_stable_detections(raw_humans, "human")

    human_machine_assignments = assign_humans_to_machines_efficiency(humans, machines, occupancy_threshold_val)

    analysis_data = {}
    for human_track_id, assigned_machine_name in human_machine_assignments.items():
        efficiency_status = None
        for machine in machines:
            if machine["class"] == assigned_machine_name:
                track_id = machine.get("track_id")
                if track_id in laplacian_history:
                    machine_crop = image[int(machine["box"][1]):int(machine["box"][3]), int(machine["box"][0]):int(machine["box"][2])]
                    if machine_crop.size > 0:
                         efficiency_status = calculate_occupancy_score(machine_crop, laplacian_history[track_id], occupancy_threshold_val)
                         exercise_tracker_eff.update_person_machine(human_track_id, assigned_machine_name, efficiency_status)


    annotated_frame, updated_laplacian_history = draw_results_with_laplacian(image, machines, humans, analysis_data, human_machine_assignments, laplacian_history, frame_idx, occupancy_threshold_val)
    return annotated_frame, updated_laplacian_history, exercise_tracker_eff.machine_efficiency

def run_streamlit_machine_efficiency_analysis(conf_threshold_eff, occupancy_iou_threshold, occupancy_threshold_val):
    cap = cv2.VideoCapture(st.session_state.video_source_efficiency)
    if not cap.isOpened():
        st.error("Error opening video source."); st.session_state.run_efficiency_analysis = False; st.rerun(); return

    if st.session_state.output_video_path_efficiency is None:
        temp_dir = tempfile.gettempdir()
        unique_filename = f"processed_video_efficiency_{int(time.time())}.mp4"
        st.session_state.output_video_path_efficiency = os.path.join(temp_dir, unique_filename)

    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(st.session_state.output_video_path_efficiency, fourcc, cap.get(cv2.CAP_PROP_FPS), (int(cap.get(cv2.CAP_PROP_FRAME_WIDTH)), int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))))

    image_placeholder = st.empty()
    st.divider()
    st.subheader("📊 Live Machine Efficiency Metrics")
    col1, col2 = st.columns(2)
    metric1_placeholder = col1.empty()
    metric2_placeholder = col2.empty()
    
    exercise_tracker_eff = ExerciseTrackerEfficiency()
    stabilizer_eff = DetectionStabilizerEfficiency()
    laplacian_history = {} 
    
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    frame_idx = 0
    
    efficiency_summary = {} 
    
    progress_bar = st.progress(0)
    
    while st.session_state.run_efficiency_analysis and cap.isOpened():
        ret, frame = cap.read()
        if not ret: break
        frame_idx += 1
        
        annotated_frame, laplacian_history, current_efficiency_data = process_frame_efficiency(
            frame, exercise_tracker_eff, stabilizer_eff, laplacian_history, frame_idx, conf_threshold_eff, occupancy_iou_threshold
        )
        for person_id, machines_eff_data in current_efficiency_data.items():
            for machine_name, status in machines_eff_data.items():
                if machine_name not in efficiency_summary:
                    efficiency_summary[machine_name] = {"Efficient": 0, "Inefficient": 0, "No Activity": 0, "Total Frames": 0}
                
                efficiency_summary[machine_name]["Total Frames"] += 1
                if "Efficient" in status:
                    efficiency_summary[machine_name]["Efficient"] += 1
                elif "Inefficient" in status:
                    efficiency_summary[machine_name]["Inefficient"] += 1
                else:
                    efficiency_summary[machine_name]["No Activity"] += 1
        
        metric1_placeholder.metric(label="Detected Machines", value=len(laplacian_history))
        
        if laplacian_history:
            first_machine_id = list(laplacian_history.keys())[0]
            first_machine_buffer = laplacian_history[first_machine_id]
            
            status = "Detecting..."
            temp_crop = frame[0:10, 0:10]
            if temp_crop.size > 0:
                status = calculate_occupancy_score(temp_crop, first_machine_buffer, occupancy_threshold_val) 
            metric2_placeholder.metric(label=f"Machine Status (Sample)", value=status)
        else:
            metric2_placeholder.metric(label="Machine Status", value="Detecting...")


        image_placeholder.image(annotated_frame, channels="BGR", use_container_width=True)
        out.write(annotated_frame)
        progress_bar.progress(min(100, int((frame_idx / total_frames) * 100)))

    cap.release()
    out.release()
    
    st.session_state.efficiency_analysis_complete = True
    st.session_state.run_efficiency_analysis = False 
    efficiency_report_path = os.path.join(tempfile.gettempdir(), f"machine_efficiency_report_{int(time.time())}.xlsx")
    report_data = []
    for machine, data in efficiency_summary.items():
        total = data["Total Frames"]
        efficient_perc = (data["Efficient"] / total * 100) if total > 0 else 0
        inefficient_perc = (data["Inefficient"] / total * 100) if total > 0 else 0
        no_activity_perc = (data["No Activity"] / total * 100) if total > 0 else 0
        report_data.append({
            "Machine": machine,
            "Efficient Usage (%)": f"{efficient_perc:.1f}%",
            "Inefficient Usage (%)": f"{inefficient_perc:.1f}%",
            "No Activity (%)": f"{no_activity_perc:.1f}%",
            "Total Frames Monitored": total
        })
    
    if report_data:
        efficiency_df = pd.DataFrame(report_data)
        try:
            with pd.ExcelWriter(efficiency_report_path, engine='openpyxl') as writer:
                efficiency_df.to_excel(writer, sheet_name='Machine Efficiency Summary', index=False)
            st.session_state.efficiency_excel_report_path = efficiency_report_path
        except Exception as e:
            st.error(f"Error saving efficiency report: {e}")
            st.session_state.efficiency_excel_report_path = None
    else:
        st.session_state.efficiency_excel_report_path = None
        
    st.rerun()

def render_machine_efficiency_analysis_page():
    st.header("⚙️ Machine Efficiency Analysis")
    st.write("Upload a video to analyze how efficiently gym machines are being utilized.")
    conf_threshold_eff = st.slider("Detection Confidence", 0.1, 1.0, st.session_state.get('conf_threshold_eff', 0.4), 0.05)
    occupancy_iou_threshold = st.slider("Occupancy IOU Threshold", 0.05, 1.0, st.session_state.get('occupancy_iou_threshold', 0.1), 0.05)
    occupancy_threshold_val = st.slider("Activity Variance Threshold", 0.1, 30.0, st.session_state.get('occupancy_threshold_val', 2.0), 0.1)
    st.info("Adjust activity variance threshold: higher values indicate more required motion for 'efficient' status.")

    st.session_state.conf_threshold_eff = conf_threshold_eff
    st.session_state.occupancy_iou_threshold = occupancy_iou_threshold
    st.session_state.occupancy_threshold_val = occupancy_threshold_val
    
    uploaded_file = st.file_uploader("Upload a video for efficiency analysis", type=["mp4", "mov", "avi", "mkv"], key="efficiency_video_uploader")

    if uploaded_file is not None:
        temp_video_path = os.path.join(tempfile.gettempdir(), uploaded_file.name)
        with open(temp_video_path, "wb") as f:
            f.write(uploaded_file.read())
        
        st.video(temp_video_path, format="video/mp4", start_time=0)
        
        if st.button("Start Efficiency Analysis", key="start_efficiency_analysis"):
            st.session_state.run_efficiency_analysis = True
            st.session_state.video_source_efficiency = temp_video_path
            st.session_state.output_video_path_efficiency = None 
            st.session_state.efficiency_analysis_complete = False
            st.session_state.efficiency_excel_report_path = None
            st.rerun()

    if st.session_state.get('run_efficiency_analysis', False):
        run_streamlit_machine_efficiency_analysis(
            st.session_state.conf_threshold_eff, 
            st.session_state.occupancy_iou_threshold,
            st.session_state.occupancy_threshold_val
        )

    if st.session_state.get('efficiency_analysis_complete', False):
        st.success("✅ Machine Efficiency Analysis Complete!")
        if st.session_state.get('output_video_path_efficiency'):
            st.video(st.session_state.output_video_path_efficiency)
            st.download_button(
                label="Download Processed Efficiency Video",
                data=open(st.session_state.output_video_path_efficiency, "rb").read(),
                file_name=os.path.basename(st.session_state.output_video_path_efficiency),
                mime="video/mp4"
            )
        if st.session_state.get('efficiency_excel_report_path'):
            st.download_button(
                label="Download Machine Efficiency Report (Excel)",
                data=open(st.session_state.efficiency_excel_report_path, "rb").read(),
                file_name=os.path.basename(st.session_state.efficiency_excel_report_path),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        st.info("The Excel report provides a summary of machine efficiency percentages.")
if 'video_source' not in st.session_state: st.session_state.video_source = None
if 'output_video_path' not in st.session_state: st.session_state.output_video_path = None
if 'run_analysis' not in st.session_state: st.session_state.run_analysis = False
if 'person_exercise_states' not in st.session_state: st.session_state.person_exercise_states = {}
if 'session_stats' not in st.session_state: st.session_state.session_stats = {}
if 'analysis_complete' not in st.session_state: st.session_state.analysis_complete = False
if 'excel_report_path' not in st.session_state: st.session_state.excel_report_path = None
if 'graph_fig' not in st.session_state: st.session_state.graph_fig = None
if 'run_non_machine_analysis' not in st.session_state: st.session_state.run_non_machine_analysis = False
if 'non_machine_video_source' not in st.session_state: st.session_state.non_machine_video_source = None
if 'non_machine_output_path' not in st.session_state: st.session_state.non_machine_output_path = None
if 'non_machine_analysis_complete' not in st.session_state: st.session_state.non_machine_analysis_complete = False
if 'non_machine_report' not in st.session_state: st.session_state.non_machine_report = None
if 'run_efficiency_analysis' not in st.session_state: st.session_state.run_efficiency_analysis = False
if 'video_source_efficiency' not in st.session_state: st.session_state.video_source_efficiency = None
if 'output_video_path_efficiency' not in st.session_state: st.session_state.output_video_path_efficiency = None
if 'efficiency_analysis_complete' not in st.session_state: st.session_state.efficiency_analysis_complete = False
if 'efficiency_excel_report_path' not in st.session_state: st.session_state.efficiency_excel_report_path = None
def render_home_page():
    st.markdown(
        """
        <style>
        .stApp {
            background-color: #0e1117;
            color: #fafafa;
        }
        .css-1d391kg {
            padding-top: 35px;
        }
        .st-emotion-cache-z5fcl4 {
            font-size: 1.1em;
        }
        .st-emotion-cache-1j00m5z {
            color: #FF4B4B;
            text-align: center;
            font-size: 3em;
            font-weight: bold;
        }
        .st-emotion-cache-vdgyx6 {
            color: #fafafa;
            text-align: center;
            font-size: 1.8em;
            margin-bottom: 25px;
        }
        .st-emotion-cache-1cpxoar a {
            color: #1E90FF !important;
        }
        .stMarkdown {
            text-align: center;
        }
        .stImage {
            border-radius: 10px;
            box-shadow: 0 4px 8px 0 rgba(0, 0, 0, 0.2);
        }
        .big-font {
            font-size:20px !important;
            font-weight: bold;
        }
        .center-text {
            text-align: center;
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    st.markdown("<h1 class='st-emotion-cache-1j00m5z'>🏋️ AI Gym Vision 🚀</h1>", unsafe_allow_html=True)
    st.markdown("<h2 class='st-emotion-cache-vdgyx6'>Your Personal AI-Powered Fitness Coach</h2>", unsafe_allow_html=True)

    st.write("---")

    col1, col2 = st.columns([1, 2]) 

    with col1:
        st.image("https://cdn.pixabay.com/photo/2020/04/09/16/55/gym-5022285_1280.jpg", 
                 caption="Transforming Your Workout Experience", use_container_width=True)
    
    with col2:
        st.markdown(
            """
            <div class='center-text'>
            <p class='big-font'>Welcome to AI Gym Vision! This application leverages advanced computer vision and AI models to provide real-time analysis and feedback on your exercises.</p>
            <p>Whether you're using gym machines or performing free-weight/bodyweight exercises, we've got you covered. Get insights into your form, track your repetitions, and understand your machine usage efficiency.</p>
            <p>Select an option from the sidebar to begin your intelligent workout journey!</p>
            </div>
            """, unsafe_allow_html=True
        )
    
    st.write("---")

    st.markdown("### ✨ Key Features:")
    features_col1, features_col2, features_col3 = st.columns(3)
    with features_col1:
        st.markdown("#### 💪 Real-time Form Analysis")
        st.info("Receive instant feedback on your exercise posture for optimal performance and injury prevention.")
    with features_col2:
        st.markdown("#### ⚙️ Machine Usage Efficiency")
        st.info("Analyze how effectively gym machines are being utilized over time, ensuring maximum productivity.")
    with features_col3:
        st.markdown("#### 📈 Progress Tracking & Reports")
        st.info("Generate detailed session reports and visualize your progress with insightful graphs.")
    
    st.markdown("---")
    st.markdown("<p class='center-text'>Developed with ❤️ using Streamlit, YOLO, and MediaPipe.</p>", unsafe_allow_html=True)
exercise_type = st.sidebar.radio("Select Analysis Type", ["Home", "Machine Exercises", "Non-Machine Exercises", "Machine Efficiency Analysis"], key="exercise_type")

if exercise_type == "Home":
    render_home_page()
elif exercise_type == "Machine Exercises":
    machine_page = st.sidebar.radio("Machine Exercise Navigation", ["Live Analysis", "Session Report"], key="machine_sub_nav")
    
    if "machine_analysis_state" not in st.session_state:
        st.session_state.machine_analysis_state = {
            'person_conf': 0.6,
            'equipment_conf': 0.6,
            'iou_threshold': 0.15
        }

    if machine_page == "Live Analysis":
        st.header("🏋️ Live Machine Exercise Analysis")
        st.write("Upload a video of your machine exercises for real-time form and rep counting.")

        uploaded_file = st.file_uploader("Upload a video", type=["mp4", "mov", "avi", "mkv"], key="machine_video_uploader")
        st.header("⚙️ Machine Exercise Settings")
        st.session_state.machine_analysis_state['person_conf'] = st.slider("Person Confidence", 0.1, 1.0, st.session_state.machine_analysis_state.get('person_conf', 0.6), 0.05)
        st.session_state.machine_analysis_state['equipment_conf'] = st.slider("Equipment Confidence", 0.1, 1.0, st.session_state.machine_analysis_state.get('equipment_conf', 0.6), 0.05)
        st.session_state.machine_analysis_state['iou_threshold'] = st.slider("Occupancy IOU Threshold", 0.05, 1.0, st.session_state.machine_analysis_state.get('iou_threshold', 0.15), 0.05)
        st.info("Adjust detection sensitivity. Higher values are stricter.")

        if uploaded_file is not None:
            temp_video_path = os.path.join(tempfile.gettempdir(), uploaded_file.name)
            with open(temp_video_path, "wb") as f:
                f.write(uploaded_file.read())
            
            st.video(temp_video_path, format="video/mp4", start_time=0)
            
            if st.button("Start Analysis", key="start_machine_analysis"):
                st.session_state.run_analysis = True
                st.session_state.video_source = temp_video_path
                st.session_state.output_video_path = None 
                st.session_state.person_exercise_states = {} 
                st.session_state.analysis_complete = False
                st.session_state.session_stats = {}
                st.session_state.excel_report_path = None
                st.session_state.graph_fig = None
                st.rerun()

        if st.session_state.run_analysis:
            run_streamlit_inference_machine(
                st.session_state.machine_analysis_state['person_conf'],
                st.session_state.machine_analysis_state['equipment_conf'],
                st.session_state.machine_analysis_state['iou_threshold']
            )

    elif machine_page == "Session Report":
        st.header("📈 Machine Exercise Session Report")
        if st.session_state.get('analysis_complete', False) and st.session_state.get('session_stats'):
            st.success("Analysis complete! Here's your session report.")
            
            if st.session_state.get('output_video_path'):
                st.video(st.session_state.output_video_path)
                with open(st.session_state.output_video_path, "rb") as video_file:
                    st.download_button(
                        label="Download Processed Video",
                        data=video_file.read(),
                        file_name=os.path.basename(st.session_state.output_video_path),
                        mime="video/mp4"
                    )

            stats = st.session_state.session_stats
            st.subheader("Summary")
            col_s1, col_s2, col_s3 = st.columns(3)
            with col_s1: st.metric(label="Total Reps Across All Exercises", value=stats['total_reps'])
            with col_s2: st.metric(label="Average Form Score", value=f"{stats['average_form_score']:.1f}")
            with col_s3: st.metric(label="Processing Time (seconds)", value=f"{stats['processing_time']:.2f}")

            st.subheader("Exercises Performed")
            if stats['exercises_performed']:
                st.write(", ".join(stats['exercises_performed']))
            else:
                st.write("No exercises performed or detected in the last session.")
            
            if st.session_state.get('excel_report_path'):
                with open(st.session_state.excel_report_path, "rb") as excel_file:
                    st.download_button(
                        label="Download Detailed Analysis Report (Excel)",
                        data=excel_file.read(),
                        file_name=os.path.basename(st.session_state.excel_report_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            
            st.subheader("Form Score Progression")
            if st.session_state.get('graph_fig'):
                st.pyplot(st.session_state.graph_fig)
            else:
                st.info("No sufficient data to generate form score graph. Perform a live analysis first.")

            st.subheader("Detailed Repetition Quality by Person")
            if st.session_state.person_exercise_states:
                detailed_data = []
                for pid, state in st.session_state.person_exercise_states.items():
                    if state.reps > 0:
                        detailed_data.append({
                            "Person ID": pid,
                            "Exercise": state.exercise_name,
                            "Reps": state.reps,
                            "Overall Quality": state.posture_metrics.quality.value,
                            "Overall Score": f"{state.posture_metrics.overall_score:.1f}",
                            "Avg Rep Processing Duration (s)": f"{np.mean(list(state.rep_durations)):.2f}" if state.rep_durations else "N/A"
                        })
                if detailed_data:
                    st.dataframe(pd.DataFrame(detailed_data))
                else:
                    st.info("No detailed exercise data recorded for any person.")
            else:
                st.info("No detailed exercise data available. Run a live analysis session.")
        else:
            st.info("No analysis has been run yet. Please go to 'Live Analysis' to start a session.")

elif exercise_type == "Non-Machine Exercises":
    render_non_machine_analysis_page()
elif exercise_type == "Machine Efficiency Analysis":
    render_machine_efficiency_analysis_page()